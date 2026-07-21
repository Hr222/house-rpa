# -*- coding: utf-8 -*-
"""房产评估对比测试：逐条调用 RPA 询价接口，比较询价结果与评估单价的偏差。

用法：
  1. 先启动 RPA 服务并确认所有平台就绪：
     python -m app.scripts.api_server --debug --manual-login
  2. 再跑本脚本：
     python test_evaluate.py

输出：results/评估对比_{timestamp}.xlsx
"""

import time
import sys
import requests
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── 配置 ──────────────────────────────────────────────
BASE_URL = "http://127.0.0.1:8000"
INPUT_FILE = "C:/Users/Administrator/Desktop/房产评估汇总表_生成2.xlsx"
OUTPUT_DIR = Path(__file__).parent / "results"
POLL_INTERVAL = 6       # 轮询间隔秒数（>5 避免连续 429）
MAX_WAIT = 600          # 单任务软等待阈值；超过后只报警，不判失败，继续阻塞等待
DEFAULT_CITY = "广州"    # Excel 无 city 列时的默认城市
REQUEST_TIMEOUT = 15    # 单次 HTTP 请求超时
READY_CHECK_INTERVAL = 5
MANUAL_BLOCK_KEYWORDS = (
    "验证码", "人机验证", "验证", "风控", "captcha", "verify",
    "登录已失效", "登录", "WAIT_MANUAL_VERIFY", "LOGIN_EXPIRED",
)


def _safe_json(resp):
    try:
        return resp.json()
    except Exception:
        return {}


def _http(method: str, path: str, **kwargs):
    kwargs.setdefault("timeout", REQUEST_TIMEOUT)
    return requests.request(method, f"{BASE_URL}{path}", **kwargs)


def _looks_like_manual_block(text: str | None) -> bool:
    if not text:
        return False
    lower = str(text).lower()
    return any(keyword.lower() in lower for keyword in MANUAL_BLOCK_KEYWORDS)


def _extract_manual_block_reasons(snapshot: dict) -> list[str]:
    reasons = []
    for platform in snapshot.get("platforms", []):
        status_code = platform.get("statusCode", "")
        message = platform.get("message", "")
        if status_code in {"WAIT_MANUAL_VERIFY", "WAIT_LOGIN"} or _looks_like_manual_block(message):
            reasons.append(f"{platform.get('name', platform.get('code', '未知平台'))}: {message or status_code}")
    return reasons


def _wait_until_service_ready(reason: str):
    """服务端进入人工验证/重新登录状态时，阻塞等待恢复。"""
    print(f"\n  ⚠ 服务端待人工处理：{reason}")
    printed_hint = False
    heartbeat = 0
    while True:
        try:
            ready_resp = _http("GET", "/health/ready")
            if ready_resp.status_code == 200:
                print("  服务已恢复就绪，继续当前任务")
                return
        except requests.RequestException as exc:
            print(f"  等待服务恢复时请求失败：{exc}")

        snapshot = {}
        try:
            status_resp = _http("GET", "/admin/status")
            snapshot = _safe_json(status_resp).get("data", {}) if status_resp.status_code == 200 else {}
        except requests.RequestException:
            snapshot = {}

        if not printed_hint:
            reasons = _extract_manual_block_reasons(snapshot)
            if reasons:
                for item in reasons:
                    print(f"  - {item}")
            print("  请在浏览器处理验证码/登录问题，并在 api_server 所在终端完成必要的回车确认。")
            printed_hint = True
        elif heartbeat % 6 == 0:
            service_code = snapshot.get("serviceStatusCode", "UNKNOWN")
            current_task = snapshot.get("currentTaskId") or "-"
            print(f"  仍在等待服务恢复... service={service_code} currentTaskId={current_task}")

        heartbeat += 1
        time.sleep(READY_CHECK_INTERVAL)


def _create_inquiry_task(city: str, community: str, area: float) -> str | None:
    """创建询价任务。服务端待人工处理时阻塞等待恢复，而不是直接失败。"""
    while True:
        try:
            resp = _http(
                "POST",
                "/inquiries",
                json={
                    "city": city,
                    "communityName": community,
                    "area": area,
                    "algorithmMode": "default",
                },
            )
        except requests.RequestException as exc:
            print(f"  创建任务请求失败，{READY_CHECK_INTERVAL}s 后重试：{exc}", flush=True)
            time.sleep(READY_CHECK_INTERVAL)
            continue

        if resp.status_code == 202:
            return _safe_json(resp).get("data", {}).get("taskId")

        if resp.status_code == 503:
            payload = _safe_json(resp)
            snapshot = payload.get("data", {})
            reason = payload.get("message") or snapshot.get("serviceStatus") or "RPA 服务未就绪"
            _wait_until_service_ready(reason)
            continue

        print(f"  ❌ 创建任务失败: {resp.status_code} {resp.text[:100]}")
        return None


def _poll_task_until_done(task_id: str) -> tuple[str, dict | None]:
    """轮询任务结果。遇到人工验证/登录短路时，返回 RETRY 让上层重试当前记录。"""
    print(f"  taskId={task_id[:12]}... 等待中", end="", flush=True)
    elapsed = 0
    warned_long_wait = False

    while True:
        time.sleep(POLL_INTERVAL)
        elapsed += POLL_INTERVAL

        try:
            resp = _http("GET", f"/inquiries/{task_id}")
        except requests.RequestException as exc:
            print(f"\n  查询任务失败，继续等待：{exc}")
            continue

        if resp.status_code == 429:
            retry = _safe_json(resp).get("data", {}).get("retryAfter", 10)
            time.sleep(retry)
            elapsed += retry
            continue

        if resp.status_code != 200:
            print(f"\n  查询任务异常: {resp.status_code} {resp.text[:100]}")
            continue

        body = _safe_json(resp).get("data", {})

        if "finalPrice" in body and body["finalPrice"] is not None:
            print(f"  完成 ({elapsed}s)")
            return "DONE", body

        status_code = body.get("statusCode", body.get("status", ""))
        if status_code in ("COMPLETED", "FAILED"):
            reason = body.get("note") or body.get("error") or body.get("branch") or "无数据"
            if _looks_like_manual_block(reason):
                print(f"\n  检测到人工验证/登录短路 ({elapsed}s): {reason}")
                _wait_until_service_ready(reason)
                return "RETRY", None
            print(f"  无数据 ({elapsed}s): {reason}")
            return "DONE", body

        if elapsed >= MAX_WAIT and not warned_long_wait:
            warned_long_wait = True
            print(f"\n  已等待 {elapsed}s，任务仍未完成；继续阻塞等待，可能正在人工过风控")

        print(".", end="", flush=True)

# ─── 读取评估表 ─────────────────────────────────────────
wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in.active

# 读取表头，建立列名→列号映射
col_map = {}
for c in range(1, ws_in.max_column + 1):
    h = ws_in.cell(row=1, column=c).value
    if h:
        col_map[str(h).strip()] = c

city_col = col_map.get("city")
area_col = col_map.get("面积㎡", 1)
price_col = col_map.get("评估单价", 2)
community_col = col_map.get("小区名称", 4)
last_data_col = max(col_map.values()) if col_map else 4
out_start_col = last_data_col + 1  # 对比列起始位置

if city_col:
    print(f"[city] 检测到 city 列（第 {city_col} 列），将逐行读取城市")
else:
    print(f"[city] 未检测到 city 列，使用默认城市: {DEFAULT_CITY}")

# 表头列名: city? | 面积㎡ | 评估单价 | 房产评估总值 | 小区名称 | ...
data = []
for row_idx in range(2, ws_in.max_row + 1):
    community = ws_in.cell(row=row_idx, column=community_col).value
    area = ws_in.cell(row=row_idx, column=area_col).value
    eval_price = ws_in.cell(row=row_idx, column=price_col).value
    if not community or not area or not eval_price:
        continue
    city = None
    if city_col:
        city = ws_in.cell(row=row_idx, column=city_col).value
        city = str(city).strip() if city else ""
    if not city:
        city = DEFAULT_CITY
    data.append({
        "city": city,
        "community": str(community).strip(),
        "area": float(area),
        "eval_price": float(eval_price),
    })

print(f"读取到 {len(data)} 条评估记录")

# ─── 检查服务就绪 ───────────────────────────────────────
r = _http("GET", "/health/ready")
if r.status_code != 200:
    print("❌ RPA 服务未就绪，请先启动并确认所有平台就绪")
    sys.exit(1)
print("[就绪] 服务 OK")

# ─── 逐条询价 ───────────────────────────────────────────
results = []

for i, item in enumerate(data):
    community = item["community"]
    area = item["area"]
    eval_price = item["eval_price"]
    city = item["city"]

    print(f"\n[{i+1}/{len(data)}] {city} {community} 面积={area}㎡ 评估单价={eval_price}")
    final_data = None
    while True:
        task_id = _create_inquiry_task(city, community, area)
        if not task_id:
            results.append({
                "社区": community, "面积": area, "评估单价": eval_price,
                "询价单价": None, "差距%": None, "分支": "ERROR",
                "在售均价": None, "成交均价": None, "状态": "FAILED",
            })
            break

        outcome, payload = _poll_task_until_done(task_id)
        if outcome == "RETRY":
            print("  当前记录将在服务恢复后自动重试")
            continue

        final_data = payload
        break

    if final_data is None:
        continue

    # 计算结果
    final_price = final_data.get("finalPrice")
    branch = final_data.get("branchCode", final_data.get("branch", ""))
    quote_avg = final_data.get("quoteAvg")
    deal_avg = final_data.get("dealAvg")

    if final_price is None:
        results.append({
            "社区": community, "面积": area, "评估单价": eval_price,
            "询价单价": None, "差距%": None, "分支": branch or "NO_DATA",
            "在售均价": quote_avg, "成交均价": deal_avg, "状态": "全部平台无数据",
        })
        continue

    diff_pct = None
    if final_price and eval_price:
        diff_pct = round((final_price - eval_price) / eval_price * 100, 2)

    # 判断分支含义
    if "QUOTE" in str(branch):
        branch_display = "采用售均价"
    elif "DEAL" in str(branch):
        branch_display = "差值>10%，取成交均价"
    elif "TAKE_LOWER" in str(branch):
        branch_display = "差值≤10%，取较低值"
    elif branch == "FAILED":
        branch_display = "无数据"
    else:
        branch_display = str(branch)

    print(f"  询价={final_price} | 评估={eval_price} | 差距={diff_pct}% | {branch_display}")

    results.append({
        "社区": community,
        "面积": area,
        "评估单价": eval_price,
        "询价单价": final_price,
        "差距%": diff_pct,
        "分支": branch_display,
        "在售均价": quote_avg,
        "成交均价": deal_avg,
        "状态": "OK",
    })

# ─── 输出 Excel（基于原表追加对比列）─────────────────────
OUTPUT_DIR.mkdir(exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = OUTPUT_DIR / f"评估对比_{timestamp}.xlsx"

# 复制原表
wb_in = openpyxl.load_workbook(INPUT_FILE)
ws = wb_in.active

# 在原表右侧追加对比表头
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

add_headers = ["询价单价", "差距比例%", "偏差评级", "是否采用售均价"]
for j, h in enumerate(add_headers):
    cell = ws.cell(row=1, column=out_start_col + j, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# 补数据
for i, r in enumerate(results):
    row = i + 2

    # 询价单价
    cell = ws.cell(row=row, column=out_start_col, value=r["询价单价"])
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    if r["询价单价"]:
        cell.number_format = '#,##0.00'

    # 差距比例%
    diff = r["差距%"]
    cell = ws.cell(row=row, column=out_start_col + 1, value=diff if diff is not None else "N/A")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    if diff is not None:
        cell.number_format = '0.00"%"'

    # 偏差评级
    if diff is None:
        rating = "N/A"
    elif abs(diff) <= 5:
        rating = "偏差小（≤5%）"
    elif abs(diff) <= 10:
        rating = "偏差中等（5%~10%）"
    else:
        rating = "偏差大（>10%）"
    cell = ws.cell(row=row, column=out_start_col + 2, value=rating)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

    # 是否采用售均价
    cell = ws.cell(row=row, column=out_start_col + 3, value=r["分支"])
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

# 列宽
from openpyxl.utils import get_column_letter
for j, width in enumerate([14, 14, 16, 28]):
    col_letter = get_column_letter(out_start_col + j)
    ws.column_dimensions[col_letter].width = width

# 汇总行
summary_row = len(results) + 3
summary_col = community_col
ws.cell(row=summary_row, column=summary_col, value="汇总").font = Font(bold=True)

valid_diffs = [r["差距%"] for r in results if r["差距%"] is not None]
if valid_diffs:
    avg_diff = sum(valid_diffs) / len(valid_diffs)
    max_diff = max(valid_diffs)
    min_diff = min(valid_diffs)
    within_10 = sum(1 for d in valid_diffs if abs(d) <= 10)
    ws.cell(row=summary_row, column=out_start_col, value=f"有效: {len(valid_diffs)}/{len(results)} 条")
    ws.cell(row=summary_row + 1, column=out_start_col, value=f"平均偏差: {avg_diff:.2f}%")
    ws.cell(row=summary_row + 2, column=out_start_col, value=f"最大偏差: {max_diff:.2f}%")
    ws.cell(row=summary_row + 3, column=out_start_col, value=f"最小偏差: {min_diff:.2f}%")
    ws.cell(row=summary_row + 4, column=out_start_col, value=f"偏差≤10%: {within_10}/{len(valid_diffs)} 条")

wb_in.save(out_path)
print(f"\n✅ 结果已保存: {out_path}")
