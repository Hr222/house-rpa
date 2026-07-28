# -*- coding: utf-8 -*-
"""将配套的询价日志导出为 Excel 工作簿，用于调试分析。"""

from __future__ import annotations

import argparse
import ast
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 支持 `python app/excel/export_operation_log_excel.py ...` 直接执行。
# 否则 sys.path 只有 app/excel，可能误导入虚拟环境中的同名第三方 app 包。
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.core.algorithm import find_weighted_price_candidates
from app.core.status import (
    BRANCH_TEXT,
    OPERATION_STATUS_TEXT,
    PlatformResultStatus,
)
from app.utils.listing_dedup import ListingDeduplicationResult, deduplicate_listings

U_QUERY_CITY = "\u67e5\u8be2\u57ce\u5e02"
U_COMMUNITY = "\u5c0f\u533a"
U_AREA = "\u9762\u79ef"
U_QM = "\u33a1"
U_LISTING_COMMUNITY = "\u5c0f\u533a\u540d\u79f0"
U_TITLE = "\u6807\u9898"
U_LAYOUT = "\u51e0\u623f\u51e0\u5385"
U_SELL = "\u552e\u4ef7"
U_TOTAL = "\u603b\u4ef7"
U_HOUSE_ID = "\u623f\u6e90\u7f16\u53f7"
U_STATUS = "\u72b6\u6001"
U_REASON = "\u539f\u56e0"
U_DEDUP_STATUS = "\u53bb\u91cd\u72b6\u6001"
U_DEAL = "\u6210\u4ea4"
U_DATE = "\u65e5\u671f"
U_COMPLETED = "\u91c7\u96c6\u5b8c\u6210"
U_QUOTE_AVG = "\u5728\u552e\u5747\u4ef7"
U_DEAL_AVG = "\u6210\u4ea4\u5747\u4ef7"
U_FINAL_PRICE = "\u6700\u7ec8\u53d6\u503c"
U_SUCCESS = "\u6210\u529f"
U_LISTING_SECTION = "\u5728\u552e\u623f\u6e90\u660e\u7ec6"
U_DEAL_SECTION = "\u6210\u4ea4\u660e\u7ec6\u4e0e\u8bf4\u660e"
U_PRICE_UNIT = "\u5143/\u5e73"
U_PRICE_UNIT_M2 = "\u5143/\u33a1"
U_WAN = "\u4e07"
U_NONE = "\u672a\u91c7\u96c6\u5230"
U_RAW = "\u539f\u59cb\u8bf4\u660e"
U_DEAL_RECORD = "\u6210\u4ea4\u8bb0\u5f55"
U_DEAL_NOTE = "\u8bf4\u660e"
U_PLATFORM = "\u5e73\u53f0"
U_CITY = "\u57ce\u5e02"
U_REQUEST_AREA = "\u8bf7\u6c42\u9762\u79ef(\u33a1)"
U_REQUEST_ID = "requestId"
U_ALGO = "algorithmMode"
U_STARTED = "startedAt"
U_FINISHED = "finishedAt"
U_ELAPSED = "elapsedSeconds"
U_BRANCH = "branchCode"
U_BRANCH_TEXT = "branch"
U_QUOTE = "quoteAvg"
U_DEALV = "dealAvg"
U_FINAL = "finalPrice"
U_NOTE = "\u5907\u6ce8"
U_UNIT_LABEL = "\u5355\u4f4d:\u5143/\u5e73"
U_ELAPSED_WORD = "\u8017\u65f6"
U_SECONDS = "\u79d2"
U_PINGMI = "\u5e73\u7c73"
U_DANJIA = "\u5355\u4ef7"
U_WU_L = "\u65e0\uff08"
U_RPAREN = "\uff09"
U_TOTAL_PREFIX = "\uff08\u5171"
U_TOTAL_SUFFIX = "\u6761\uff09"
U_LOG_INCOMPLETE = "\u65e5\u5fd7\u4e2d\u672a\u89e3\u6790\u5230\u5b8c\u6574\u7ed3\u679c"
U_NO_USABLE_DATA = "\u672a\u6293\u53d6\u5230\u53ef\u7528\u4e8e\u8ba1\u7b97\u7684\u6709\u6548\u6570\u636e\uff0c\u6700\u7ec8\u65e0\u53ef\u7528\u62a5\u4ef7"
U_STATUS_SUCCESS = PlatformResultStatus.SUCCESS.value
U_PINGMI_LABEL = "\u5e73\u7c73"
U_FAILED = "FAILED"
U_WEIGHTED_MEDIAN = "WEIGHTED_MEDIAN"
U_WEIGHTED_MEDIAN_COMBINED = "WEIGHTED_MEDIAN_COMBINED"
U_WEAK_REFERENCE = "弱参考"

# Excel 展示用中文。上面的 U_* 常量同时参与日志解析，不能直接改名或改值。
D_REQUEST_ID = "请求编号"
D_ALGORITHM_MODE = "算法模式"
D_STARTED = "开始时间"
D_FINISHED = "完成时间"
D_ELAPSED = "耗时(秒)"
D_RESULT = "结果"
D_QUOTE = "在售均价"
D_DEAL = "成交均价"
D_FINAL = "最终取值"
D_BRANCH = "决策分支"
D_BRANCH_TEXT = "分支说明"
D_CANDIDATES = "当前算法候选"
D_STRICT_CANDIDATES = "严格±1㎡候选"
D_WEAK_REFERENCE = "弱参考信息"

ALGORITHM_MODE_TEXT = {
    "DEFAULT": "加权落点中位数算法",
}
ALGORITHM_DESCRIPTIONS = {
    "DEFAULT": (
        "加权落点中位数算法：汇集所有平台的有效房源落点，每条房源按出现次数计票；低频孤立峰过滤，"
        "主峰明确时返回一个中位数并打折，多峰时选择最低价格峰中位数直接返回，不打折；"
        "存在真实目标面积成交价时，再将挂牌结果与成交结果做等权平均；没有符合目标面积的成交价时沿用挂牌结果。"
    ),
}

ENCODINGS = ("utf-8", "utf-8-sig", "gb18030")
INVALID_SHEET_CHARS = set('[]:*?/\\')


@dataclass
class ListingRow:
    platform: str
    community_name: str
    title: str
    area: Optional[float]
    layout: str
    unit_price: Optional[float]
    total_price: Optional[float]
    house_id: str = ""


@dataclass
class DealRow:
    platform: str
    area: Optional[float]
    date: str
    total_price: Optional[float]
    price: Optional[float]


@dataclass
class InquiryRecord:
    started_at: str
    city: str
    community_name: str
    area: Optional[float]
    request_id: Optional[str] = None
    algorithm_mode: str = "DEFAULT"
    elapsed_seconds: Optional[float] = None
    finished_at: Optional[str] = None
    quote_avg: Optional[float] = None
    deal_avg: Optional[float] = None
    final_price: Optional[float] = None
    final_price_logged: bool = False
    success: bool = False
    branch_code: Optional[str] = None
    branch_text: Optional[str] = None
    branch_code_logged: bool = False
    reference_code: Optional[str] = None
    reference_area_tolerance: Optional[str] = None
    reference_area_min: Optional[str] = None
    reference_area_max: Optional[str] = None
    reference_listing_count: Optional[str] = None
    listings: list[ListingRow] = field(default_factory=list)
    deals: list[DealRow] = field(default_factory=list)
    platform_notes: dict[str, dict[str, str]] = field(default_factory=dict)


@dataclass(frozen=True)
class EvaluationRow:
    """用于解释对应日志结果的一行评估数据。"""

    row_number: int
    city: str
    area: Optional[float]
    community_name: str
    evaluation_price: Optional[float]


@dataclass
class AnalysisRow:
    """分析汇总表中展示的对比和归因结果。"""

    evaluation: EvaluationRow
    record: Optional[InquiryRecord]
    raw_peak_price: Optional[float] = None
    nearest_peak_diff: Optional[float] = None
    raw_diff: Optional[float] = None
    final_diff: Optional[float] = None
    candidate_text: str = ""
    strict_candidate_text: str = ""
    weak_reference_text: str = ""
    conclusion: str = ""
    evaluation_scope: str = ""
    strict_listing_count: int = 0
    weak_listing_count: int = 0
    raw_listing_count: int = 0
    same_platform_listing_count: int = 0
    deduplicated_listing_count: int = 0
    cross_platform_duplicate_text: str = ""


def _listing_deduplication(record: InquiryRecord) -> ListingDeduplicationResult[ListingRow]:
    """执行同平台及保守的跨平台去重。"""
    return deduplicate_listings(record.listings)


def _deduplicated_listing_rows(record: InquiryRecord) -> list[ListingRow]:
    """返回参与跨平台价格频次统计的数据行。"""
    return list(_listing_deduplication(record).items)


def _format_cross_platform_duplicate_text(
    result: ListingDeduplicationResult[ListingRow],
) -> str:
    """格式化重复房源组，写入工作簿审计列。"""
    parts: list[str] = []
    for group in result.cross_platform_groups:
        platforms = "/".join(str(getattr(item, "platform", "")) for item in group.members)
        representative = group.representative
        parts.append(
            f"{platforms}: {representative.unit_price:,.0f}元/平, "
            f"{representative.area:g}平米, {representative.layout}"
            f" ({group.reason})"
        )
    return "; ".join(parts)


def _listing_dedup_statuses(record: InquiryRecord) -> dict[int, str]:
    """说明从价格频次统计中移除、但仍保留原始数据的行。"""
    result = _listing_deduplication(record)
    statuses: dict[int, str] = {}
    for group in result.cross_platform_groups:
        for listing in group.members[1:]:
            statuses[id(listing)] = "\u8de8\u5e73\u53f0\u91cd\u590d\uff08\u4e0d\u53c2\u4e0e\u9891\u7387\u7edf\u8ba1\uff09"

    kept_ids = {id(listing) for listing in result.items}
    for listing in record.listings:
        if id(listing) not in kept_ids:
            statuses.setdefault(
                id(listing),
                "\u540c\u5e73\u53f0\u91cd\u590d\uff08\u4e0d\u53c2\u4e0e\u9891\u7387\u7edf\u8ba1\uff09",
            )
    return statuses


def _weighted_median_candidates_for_record(record: InquiryRecord):
    """根据记录中的去重房源重新构建加权落点中位数候选值。"""
    listings = _deduplicated_listing_rows(record)
    prices = [
        float(listing.unit_price)
        for listing in listings
        if listing.unit_price is not None and listing.unit_price > 0
    ]
    return find_weighted_price_candidates([prices])


def _weighted_median_candidates_for_listings(listings: list[ListingRow]):
    prices = [
        float(listing.unit_price)
        for listing in listings
        if listing.unit_price is not None and listing.unit_price > 0
    ]
    return find_weighted_price_candidates([prices])


def _weak_reference_notes(record: InquiryRecord) -> list[tuple[str, dict[str, str]]]:
    return [
        (platform, note)
        for platform, note in record.platform_notes.items()
        if note.get("reference_code")
    ]


def _format_optional_number(value: Optional[float]) -> str:
    return f"{value:.2f}" if value is not None else "?"


def _format_weak_reference_text(record: InquiryRecord) -> str:
    parts = []
    for platform, note in _weak_reference_notes(record):
        tolerance = to_float(note.get("reference_area_tolerance"))
        area_min = to_float(note.get("reference_area_min"))
        area_max = to_float(note.get("reference_area_max"))
        count = note.get("reference_listing_count", "0")
        parts.append(
            f"{platform}: {note.get('reference_code')} "
            f"±{_format_optional_number(tolerance)}㎡ "
            f"({_format_optional_number(area_min)}~"
            f"{_format_optional_number(area_max)}㎡)，额外{count}条"
        )
    return "; ".join(parts)


def _format_platform_reason(note: dict[str, str]) -> str:
    reason = note.get("reason", "")
    if not note.get("reference_code"):
        return reason
    tolerance = to_float(note.get("reference_area_tolerance"))
    area_min = to_float(note.get("reference_area_min"))
    area_max = to_float(note.get("reference_area_max"))
    weak_text = (
        f"{note.get('reference_code')} "
        f"±{_format_optional_number(tolerance)}㎡ "
        f"({_format_optional_number(area_min)}~"
        f"{_format_optional_number(area_max)}㎡)，"
        f"额外{note.get('reference_listing_count', '0')}条"
    )
    return f"{reason}; {weak_text}" if reason else weak_text


def _format_final_weak_reference_text(record: InquiryRecord) -> str:
    """仅格式化最终结果实际使用的弱参考信息。"""
    if not record.reference_code:
        return ""
    tolerance = to_float(record.reference_area_tolerance)
    area_min = to_float(record.reference_area_min)
    area_max = to_float(record.reference_area_max)
    return (
        f"referenceCode={record.reference_code} "
        f"referenceAreaTolerance={_format_optional_number(tolerance)} "
        f"referenceAreaMin={_format_optional_number(area_min)} "
        f"referenceAreaMax={_format_optional_number(area_max)} "
        f"referenceListingCount={record.reference_listing_count or '0'}"
    )


def _format_candidate_text(candidates) -> str:
    return "; ".join(
        f"{candidate.quote_price:,.0f}（{candidate.count}条，{candidate.frequency:.2%}）"
        for candidate in candidates
    )


def read_evaluation_rows(path: Path) -> list[EvaluationRow]:
    """从评估工作簿的第一个工作表读取评估价格。"""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if not header_values:
            return []

        headers = {clean_text(value): index for index, value in enumerate(header_values)}

        def find_column(*names: str) -> Optional[int]:
            for name in names:
                if name in headers:
                    return headers[name]
            return None

        city_column = find_column("city", "城市")
        area_column = find_column("面积㎡", "面积", "请求面积(㎡)")
        community_column = find_column("小区名称", "小区")
        price_column = find_column("评估单价", "评估价", "评估价格")
        missing = [
            label
            for label, column in (
                ("城市", city_column),
                ("面积", area_column),
                ("小区名称", community_column),
                ("评估单价", price_column),
            )
            if column is None
        ]
        if missing:
            raise ValueError(f"Evaluation workbook is missing columns: {', '.join(missing)}")

        evaluation_rows: list[EvaluationRow] = []
        for row_number, values in enumerate(rows, start=2):
            community_name = clean_text(values[community_column])
            if not community_name or community_name == "汇总":
                continue
            evaluation_rows.append(
                EvaluationRow(
                    row_number=row_number,
                    city=clean_text(values[city_column]),
                    area=to_float(values[area_column]),
                    community_name=community_name,
                    evaluation_price=to_float(values[price_column]),
                )
            )
        return evaluation_rows
    finally:
        workbook.close()


def _match_record(
    evaluation: EvaluationRow,
    records: list[InquiryRecord],
    unused_indexes: set[int],
) -> Optional[InquiryRecord]:
    """匹配一行评估数据，但不合并重复出现的小区名称。"""
    matches = []
    for index in unused_indexes:
        record = records[index]
        if record.city != evaluation.city or record.community_name != evaluation.community_name:
            continue
        if evaluation.area is None or record.area is None:
            area_distance = 0.0
        else:
            area_distance = abs(record.area - evaluation.area)
            if area_distance > 0.2:
                continue
        matches.append((area_distance, index, record))
    if not matches:
        return None
    _, index, record = min(matches, key=lambda item: (item[0], item[1]))
    unused_indexes.remove(index)
    return record


def analyze_evaluation_rows(
    evaluation_rows: list[EvaluationRow],
    records: list[InquiryRecord],
) -> list[AnalysisRow]:
    """将评估价格与原始峰值和最终值进行对比。"""
    unused_indexes = set(range(len(records)))
    analysis_rows: list[AnalysisRow] = []
    for evaluation in evaluation_rows:
        record = _match_record(evaluation, records, unused_indexes)
        analysis = AnalysisRow(evaluation=evaluation, record=record)
        if record is None:
            analysis.conclusion = "日志未匹配，不参与统计"
            analysis.evaluation_scope = "未匹配"
            analysis_rows.append(analysis)
            continue

        deduplication = _listing_deduplication(record)
        analysis.raw_listing_count = deduplication.raw_count
        analysis.same_platform_listing_count = len(deduplication.same_platform_items)
        analysis.deduplicated_listing_count = len(deduplication.items)
        analysis.cross_platform_duplicate_text = _format_cross_platform_duplicate_text(
            deduplication
        )
        analysis.weak_reference_text = _format_final_weak_reference_text(record)
        weak_reference_used = bool(record.reference_code)
        deduplicated_listings = list(deduplication.items)
        strict_listings = [
            listing
            for listing in deduplicated_listings
            if (
                record.area is not None
                and listing.area is not None
                and abs(listing.area - record.area) <= 1.0
            )
        ]
        analysis.strict_listing_count = len(strict_listings)
        analysis.weak_listing_count = (
            int(record.reference_listing_count or 0) if weak_reference_used else 0
        )
        candidates = (
            _weighted_median_candidates_for_record(record)
            if record.algorithm_mode == "DEFAULT"
            else []
        )
        analysis.candidate_text = _format_candidate_text(candidates)
        strict_candidates = (
            _weighted_median_candidates_for_listings(strict_listings)
            if record.algorithm_mode == "DEFAULT"
            else []
        )
        analysis.strict_candidate_text = _format_candidate_text(strict_candidates)
        if record.final_price is None or evaluation.evaluation_price is None:
            analysis.conclusion = "无数据，不参与准确率统计"
            analysis.evaluation_scope = "无数据"
            analysis_rows.append(analysis)
            continue

        evaluation_price = evaluation.evaluation_price
        if candidates:
            selected = min(candidates, key=lambda candidate: candidate.quote_price)
            analysis.raw_peak_price = selected.quote_price
            peak_diffs = [
                abs(candidate.quote_price - evaluation_price) / evaluation_price
                for candidate in candidates
            ]
            analysis.nearest_peak_diff = min(peak_diffs)
            analysis.raw_diff = abs(selected.quote_price - evaluation_price) / evaluation_price
            analysis.final_diff = abs(record.final_price - evaluation_price) / evaluation_price
            if len(candidates) > 1:
                if analysis.nearest_peak_diff > 0.10:
                    analysis.conclusion = "原始数据/评估基准不匹配，排除评价"
                    analysis.evaluation_scope = "多峰排除"
                else:
                    analysis.conclusion = "多峰：按规则取最低峰中位数，不打折"
                    analysis.evaluation_scope = "多峰单独观察"
                analysis_rows.append(analysis)
                continue
            raw_peak = selected.quote_price
        else:
            raw_peak = record.quote_avg
            analysis.raw_peak_price = raw_peak
            if raw_peak is None:
                analysis.conclusion = "无数据，不参与准确率统计"
                analysis.evaluation_scope = "无数据"
                analysis_rows.append(analysis)
                continue
            analysis.raw_diff = abs(raw_peak - evaluation_price) / evaluation_price
            analysis.final_diff = abs(record.final_price - evaluation_price) / evaluation_price

        if analysis.raw_diff is None:
            analysis.conclusion = "无数据，不参与准确率统计"
            analysis.evaluation_scope = "无数据"
        elif analysis.raw_diff > 0.10:
            analysis.conclusion = "原始数据偏高/偏低，不是算法问题"
            analysis.evaluation_scope = "原始数据排除"
        elif record.deal_avg is None:
            analysis.conclusion = "无真实成交价，按规则九折"
            analysis.evaluation_scope = "计入单峰评价"
        elif analysis.final_diff is not None and analysis.final_diff > 0.10:
            analysis.conclusion = "算法处理造成偏差"
            analysis.evaluation_scope = "算法问题"
        else:
            analysis.conclusion = "算法结果在阈值内"
            analysis.evaluation_scope = "计入单峰评价"
        analysis_rows.append(analysis)
    return analysis_rows


PREFIX_RE = re.compile(r"^(?P<ts>\S+ \S+) \[(?P<level>\w+)\] (?P<logger>[^ ]+) - (?P<msg>.*)$")
START_RE = re.compile(
    rf"^{re.escape(U_QUERY_CITY)}: (?P<city>.*?), {re.escape(U_COMMUNITY)}: (?P<community>.*?), "
    rf"{re.escape(U_AREA)}: (?P<area>[\d.]+){re.escape(U_QM)}$"
)
SUMMARY_RE = re.compile(
    rf"^(?P<label>{re.escape(U_QUOTE_AVG)}|{re.escape(U_DEAL_AVG)}|{re.escape(U_FINAL_PRICE)})"
    rf"\({re.escape(U_UNIT_LABEL)}\): (?P<value>.+)$"
)
FINAL_WEAK_REFERENCE_RE = re.compile(
    r"^finalWeakReference: referenceCode=(?P<code>\S+) "
    r"referenceAreaTolerance=(?P<tolerance>-?[\d.]+) "
    r"referenceAreaMin=(?P<area_min>-?[\d.]+) "
    r"referenceAreaMax=(?P<area_max>-?[\d.]+) "
    r"referenceListingCount=(?P<count>\d+)$"
)
FINAL_BRANCH_RE = re.compile(r"^finalBranch: branchCode=(?P<branch>\S+)$")
WEAK_REFERENCE_RE = re.compile(
    rf"^(?P<platform>[^:]+){re.escape(U_WEAK_REFERENCE)}: "
    r"referenceCode=(?P<code>\S+) "
    r"referenceAreaTolerance=(?P<tolerance>-?[\d.]+) "
    r"referenceAreaMin=(?P<area_min>-?[\d.]+) "
    r"referenceAreaMax=(?P<area_max>-?[\d.]+) "
    r"referenceListingCount=(?P<count>\d+)$"
)
RUNTIME_RE = re.compile(
    rf"^\[{re.escape(U_COMPLETED)}\] request=(?P<request>\{{.*\}}) "
    rf"{re.escape(U_ELAPSED_WORD)}=(?P<elapsed>[\d.]+){re.escape(U_SECONDS)}"
)
LISTING_RE = re.compile(
    r"^(?P<platform>[^:]+): "
    r"\{" + re.escape(U_LISTING_COMMUNITY) + r": (?P<community_name>.*?), "
    + re.escape(U_TITLE) + r": (?P<title>.*?), "
    + re.escape(U_AREA) + r": (?P<area>.*?)" + re.escape(U_PINGMI) + r", "
    + re.escape(U_LAYOUT) + r": (?P<layout>.*?), "
    + re.escape(U_SELL) + r": (?P<unit_price>.*?)" + re.escape(U_PRICE_UNIT) + r", "
    + re.escape(U_TOTAL) + r": (?P<total_price>.*?)" + re.escape(U_WAN)
    + rf"(?:, {re.escape(U_HOUSE_ID)}: (?P<house_id>.*?))?\}}$"
)
NO_DATA_RE = re.compile(
    rf"^(?P<platform>[^:]+): \{{{re.escape(U_STATUS)}: (?P<status>.*?), {re.escape(U_REASON)}: (?P<reason>.*)\}}$"
)
DEAL_RECORD_RE = re.compile(
    r"^(?P<platform>.+?)" + re.escape(U_DEAL) + r": "
    r"\{" + re.escape(U_AREA) + r": (?P<area>.*?)" + re.escape(U_QM) + r", "
    + re.escape(U_DATE) + r": (?P<date>.*?), "
    + re.escape(U_TOTAL) + r": (?P<total_price>.*?)" + re.escape(U_WAN) + r", "
    + re.escape(U_DANJIA) + r": (?P<price>.*?)" + re.escape(U_PRICE_UNIT) + r"\}$"
)
DEAL_SOURCE_RE = re.compile(
    r"^(?P<platform>.+?)" + re.escape(U_DEAL) + r": "
    + re.escape(U_WU_L) + r"(?P<source>.*?)(?: (?P<price>.*?))?"
    + re.escape(U_PRICE_UNIT_M2) + re.escape(U_RPAREN) + r"$"
)
DEAL_NONE_RE = re.compile(r"^(?P<platform>.+?)" + re.escape(U_DEAL) + r": " + re.escape(U_NONE) + r"$")
DEAL_LIST_RE = re.compile(
    r"^(?P<platform>.+?)" + re.escape(U_DEAL) + r": \[(?P<prices>.*?)\]"
    + re.escape(U_TOTAL_PREFIX) + r"(?P<count>\d+)" + re.escape(U_TOTAL_SUFFIX) + r"$"
)


def read_log_lines(path: Path) -> list[str]:
    last_exc: Optional[Exception] = None
    for encoding in ENCODINGS:
        try:
            return path.read_text(encoding=encoding).splitlines()
        except Exception as exc:
            last_exc = exc
    assert last_exc is not None
    raise last_exc


def clean_text(value: object | None) -> str:
    return "" if value is None else str(value).strip()


def to_float(value: Optional[str]) -> Optional[float]:
    text = clean_text(value).replace(",", "")
    if not text or text in {"-", "\u2014", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def display_algorithm_mode(mode: Optional[str]) -> str:
    """返回 Excel 中使用的算法模式中文标签。"""
    return ALGORITHM_MODE_TEXT.get(mode or "DEFAULT", "未知算法")


def algorithm_description(mode: Optional[str]) -> str:
    """返回工作簿备注中使用的取值规则说明。"""
    return ALGORITHM_DESCRIPTIONS.get(mode or "DEFAULT", "未识别的算法取值规则。")


def display_status(status: Optional[str]) -> str:
    """在 Excel 展示层将状态码转换为中文。"""
    if not status:
        return ""
    return OPERATION_STATUS_TEXT.get(str(status), "状态异常")


def display_branch(branch: Optional[str]) -> str:
    """在 Excel 展示层将决策分支码转换为中文。"""
    if not branch:
        return ""
    return BRANCH_TEXT.get(str(branch), "其他处理")


def has_captured_data(record: InquiryRecord) -> bool:
    """判断日志中是否存在平台返回数据的证据。"""
    if record.listings or record.deals or record.quote_avg is not None or record.deal_avg is not None:
        return True
    return any(
        note.get("status") == U_STATUS_SUCCESS
        or note.get("deal_source")
        or note.get("deal_prices_text")
        for note in record.platform_notes.values()
    )


def calculation_no_data_description(mode: Optional[str]) -> str:
    """说明已采集数据为何未产生最终报价。"""
    return (
        "已抓到在售数据，但价格落点未形成符合频次要求的有效候选，或每条价格相对落点中心价的偏差超过10%，"
        "无法形成明确落点，最终无可用报价。"
    )


def weighted_median_no_data_detail(record: InquiryRecord) -> str:
    """说明加权价格簇的覆盖率或偏差为何未达标。"""
    deduplicated_listings = _deduplicated_listing_rows(record)
    prices = []
    for listing in deduplicated_listings:
        if listing.unit_price is not None and listing.unit_price > 0:
            prices.append(float(listing.unit_price))

    diagnostic = _diagnose_weighted_median_quote([prices])
    if diagnostic is None:
        return "没有足够的有效在售价格形成候选价格簇。"

    unique_prices = sorted(set(diagnostic.prices))
    if len(unique_prices) <= 6:
        price_text = "、".join(f"{price:.0f}" for price in unique_prices)
        candidate_text = f"候选价格为 {price_text}"
    else:
        candidate_text = (
            f"候选价格范围为 {unique_prices[0]:.0f}～{unique_prices[-1]:.0f}"
            f"（共{len(diagnostic.prices)}条）"
        )

    coverage_text = f"覆盖{diagnostic.coverage * 100:.2f}%权重"
    deviation_text = f"最大单条偏差{diagnostic.max_relative_deviation * 100:.2f}%"
    if diagnostic.coverage < 0.60:
        return (
            f"{candidate_text}，中心价约{diagnostic.center:.0f}，{coverage_text}，低于60%权重要求；"
            f"虽然{deviation_text}不超过10%，但覆盖不足，最终无可用报价。"
        )
    return (
        f"最接近的60%权重候选价格簇{candidate_text}，中心价约{diagnostic.center:.0f}，"
        f"{coverage_text}，但{deviation_text}超过10%，最终无可用报价。"
    )


def weighted_median_candidates_detail(record: InquiryRecord) -> str:
    """结合所有平台说明当前基于频次的候选价格。"""
    candidates = _weighted_median_candidates_for_record(record)
    if not candidates:
        return "没有足够的有效在售价格形成候选价格簇。"

    parts = [
        (
            f"中位数{candidate.quote_price:.0f}，"
            f"{'直接返回' if len(candidates) > 1 else '折后' }"
            f"{candidate.quote_price if len(candidates) > 1 else candidate.final_price:.0f}，"
            f"{candidate.count}条，占{candidate.frequency * 100:.2f}%"
        )
        for candidate in candidates
    ]
    if len(candidates) > 1:
        return "检测到多个高频价格落点，分别输出候选：" + "；".join(parts)
    return "检测到明确主峰：" + parts[0]


@dataclass(frozen=True)
class WeightedDiagnostic:
    prices: tuple[float, ...]
    center: float
    coverage: float
    max_relative_deviation: float


def _diagnose_weighted_median_quote(
    price_groups: list[list[float]],
    min_coverage: float = 0.60,
    max_relative_deviation: float = 0.10,
) -> Optional[WeightedDiagnostic]:
    """不导入 app 包，直接诊断加权价格簇。"""
    weighted_prices: list[tuple[float, float]] = []
    for group in price_groups:
        valid = [float(price) for price in group if price is not None and price > 0]
        if valid:
            weight = 1.0 / len(valid)
            weighted_prices.extend((price, weight) for price in valid)
    weighted_prices.sort(key=lambda item: item[0])
    if not weighted_prices:
        return None

    total_weight = sum(weight for _, weight in weighted_prices)
    target_weight = total_weight * min_coverage

    def weighted_median(items: list[tuple[float, float]]) -> Optional[float]:
        if not items:
            return None
        half = sum(weight for _, weight in items) / 2
        cumulative = 0.0
        for index, (price, weight) in enumerate(items):
            cumulative += weight
            if cumulative > half:
                return price
            if abs(cumulative - half) <= 1e-12 and index + 1 < len(items):
                return (price + items[index + 1][0]) / 2
        return items[-1][0]

    best_valid: Optional[tuple[float, float, int, int, int, float]] = None
    for start in range(len(weighted_prices)):
        interval_weight = 0.0
        for end in range(start, len(weighted_prices)):
            interval_weight += weighted_prices[end][1]
            selected = weighted_prices[start : end + 1]
            center = weighted_median(selected)
            if center is None or center <= 0:
                continue
            deviation = max(abs(price - center) / center for price, _ in selected)
            if deviation > max_relative_deviation:
                continue
            candidate = (interval_weight, -deviation, -(end - start), start, end, center)
            if best_valid is None or candidate > best_valid:
                best_valid = candidate

    if best_valid is not None:
        interval_weight, negative_deviation, _, start, end, center = best_valid
        if interval_weight < target_weight:
            selected = weighted_prices[start : end + 1]
            return WeightedDiagnostic(
                prices=tuple(price for price, _ in selected),
                center=center,
                coverage=interval_weight / total_weight,
                max_relative_deviation=-negative_deviation,
            )

    # 没有有效价格簇达到目标：展示最窄的目标大小区间，便于工作簿解释被拒绝的数据。
    closest: Optional[tuple[float, float, int, int, int, float]] = None
    for start in range(len(weighted_prices)):
        interval_weight = 0.0
        for end in range(start, len(weighted_prices)):
            interval_weight += weighted_prices[end][1]
            if interval_weight < target_weight:
                continue
            selected = weighted_prices[start : end + 1]
            center = weighted_median(selected)
            if center is None or center <= 0:
                continue
            deviation = max(abs(price - center) / center for price, _ in selected)
            candidate = (deviation, -interval_weight, end - start, start, end, center)
            if closest is None or candidate < closest:
                closest = candidate
            break
    if closest is None:
        return None
    deviation, negative_weight, _, start, end, center = closest
    selected = weighted_prices[start : end + 1]
    return WeightedDiagnostic(
        prices=tuple(price for price, _ in selected),
        center=center,
        coverage=(-negative_weight) / total_weight,
        max_relative_deviation=deviation,
    )


def infer_branch(quote_avg: Optional[float], deal_avg: Optional[float]) -> tuple[str, str]:
    if quote_avg is None:
        return U_FAILED, "\u6ca1\u6709\u660e\u786e\u7684\u4ef7\u683c\u843d\u70b9"
    if deal_avg is not None:
        return U_WEIGHTED_MEDIAN_COMBINED, "挂牌价与成交价等权平均"
    return U_WEIGHTED_MEDIAN, "\u4e3b\u8981\u4ef7\u683c\u843d\u70b9\u52a0\u6743\u4e2d\u4f4d\u6570\u6253\u6298"


def finalize_record(record: InquiryRecord) -> None:
    if record.algorithm_mode == "DEFAULT" and record.listings:
        candidates = _weighted_median_candidates_for_record(record)
        if len(candidates) == 1:
            candidate = candidates[0]
            record.quote_avg = candidate.quote_price
            record.final_price = candidate.final_price
            record.success = True
            record.branch_code = U_WEIGHTED_MEDIAN
            record.branch_text = BRANCH_TEXT[U_WEIGHTED_MEDIAN]
        elif len(candidates) > 1:
            selected = min(candidates, key=lambda candidate: candidate.quote_price)
            record.quote_avg = selected.quote_price
            record.final_price = selected.quote_price
            record.success = True
            record.branch_code = "WEIGHTED_MEDIAN_MULTI"
            record.branch_text = BRANCH_TEXT[record.branch_code]
        if record.success and record.deal_avg is not None and record.quote_avg is not None:
            # 有真实成交价时，挂牌峰值不打折，直接与成交价等权平均。
            record.final_price = (record.quote_avg + record.deal_avg) / 2
            record.branch_code = U_WEIGHTED_MEDIAN_COMBINED
            record.branch_text = BRANCH_TEXT[U_WEIGHTED_MEDIAN_COMBINED]
            return
        if record.success:
            return

    if record.branch_code_logged:
        record.success = record.branch_code in {
            U_WEIGHTED_MEDIAN,
            "WEIGHTED_MEDIAN_MULTI",
            U_WEIGHTED_MEDIAN_COMBINED,
        } and record.final_price is not None
        record.branch_text = BRANCH_TEXT.get(record.branch_code, record.branch_code)
        return

    record.success = record.final_price is not None
    record.branch_code, record.branch_text = infer_branch(record.quote_avg, record.deal_avg)


def parse_records(lines: list[str]) -> list[InquiryRecord]:
    records: list[InquiryRecord] = []
    current: Optional[InquiryRecord] = None

    for line in lines:
        prefix_match = PREFIX_RE.match(line)
        if not prefix_match:
            continue
        ts = prefix_match.group("ts")
        logger = prefix_match.group("logger")
        msg = prefix_match.group("msg")

        if logger == "app.service":
            start_match = START_RE.match(msg)
            if start_match:
                if current is not None:
                    finalize_record(current)
                    records.append(current)
                current = InquiryRecord(
                    started_at=ts,
                    city=clean_text(start_match.group("city")),
                    community_name=clean_text(start_match.group("community")),
                    area=to_float(start_match.group("area")),
                )
                continue

        if current is None:
            continue

        if logger == "app.service":
            final_branch_match = FINAL_BRANCH_RE.match(msg)
            if final_branch_match:
                current.branch_code = clean_text(final_branch_match.group("branch"))
                current.branch_text = BRANCH_TEXT.get(
                    current.branch_code, current.branch_code
                )
                current.branch_code_logged = True
                continue

            final_weak_reference_match = FINAL_WEAK_REFERENCE_RE.match(msg)
            if final_weak_reference_match:
                current.reference_code = clean_text(final_weak_reference_match.group("code"))
                current.reference_area_tolerance = clean_text(
                    final_weak_reference_match.group("tolerance")
                )
                current.reference_area_min = clean_text(
                    final_weak_reference_match.group("area_min")
                )
                current.reference_area_max = clean_text(
                    final_weak_reference_match.group("area_max")
                )
                current.reference_listing_count = clean_text(
                    final_weak_reference_match.group("count")
                )
                continue

            weak_reference_match = WEAK_REFERENCE_RE.match(msg)
            if weak_reference_match:
                platform = clean_text(weak_reference_match.group("platform"))
                current.platform_notes.setdefault(platform, {"status": U_STATUS_SUCCESS})
                current.platform_notes[platform].update(
                    {
                        "reference_code": clean_text(weak_reference_match.group("code")),
                        "reference_area_tolerance": clean_text(
                            weak_reference_match.group("tolerance")
                        ),
                        "reference_area_min": clean_text(
                            weak_reference_match.group("area_min")
                        ),
                        "reference_area_max": clean_text(
                            weak_reference_match.group("area_max")
                        ),
                        "reference_listing_count": clean_text(
                            weak_reference_match.group("count")
                        ),
                    }
                )
                continue

            listing_match = LISTING_RE.match(msg)
            if listing_match:
                platform = clean_text(listing_match.group("platform"))
                current.listings.append(
                    ListingRow(
                        platform=platform,
                        community_name=clean_text(listing_match.group("community_name")),
                        title=clean_text(listing_match.group("title")),
                        area=to_float(listing_match.group("area")),
                        layout=clean_text(listing_match.group("layout")),
                        unit_price=to_float(listing_match.group("unit_price")),
                        total_price=to_float(listing_match.group("total_price")),
                        house_id=clean_text(listing_match.group("house_id")),
                    )
                )
                current.platform_notes.setdefault(platform, {"status": U_STATUS_SUCCESS})
                continue

            no_data_match = NO_DATA_RE.match(msg)
            if no_data_match:
                current.platform_notes[clean_text(no_data_match.group("platform"))] = {
                    "status": clean_text(no_data_match.group("status")),
                    "reason": clean_text(no_data_match.group("reason")),
                }
                continue

            deal_record_match = DEAL_RECORD_RE.match(msg)
            if deal_record_match:
                platform = clean_text(deal_record_match.group("platform"))
                current.deals.append(
                    DealRow(
                        platform=platform,
                        area=to_float(deal_record_match.group("area")),
                        date=clean_text(deal_record_match.group("date")),
                        total_price=to_float(deal_record_match.group("total_price")),
                        price=to_float(deal_record_match.group("price")),
                    )
                )
                current.platform_notes.setdefault(platform, {"status": U_STATUS_SUCCESS})
                continue

            deal_source_match = DEAL_SOURCE_RE.match(msg)
            if deal_source_match:
                platform = clean_text(deal_source_match.group("platform"))
                current.platform_notes.setdefault(platform, {"status": U_STATUS_SUCCESS})
                current.platform_notes[platform].update(
                    {
                        "deal_source": clean_text(deal_source_match.group("source")),
                        "deal_price": clean_text(deal_source_match.group("price")),
                    }
                )
                continue

            deal_none_match = DEAL_NONE_RE.match(msg)
            if deal_none_match:
                platform = clean_text(deal_none_match.group("platform"))
                current.platform_notes.setdefault(platform, {"status": U_STATUS_SUCCESS})
                current.platform_notes[platform]["deal_note"] = U_NONE
                continue

            deal_list_match = DEAL_LIST_RE.match(msg)
            if deal_list_match:
                platform = clean_text(deal_list_match.group("platform"))
                current.platform_notes.setdefault(platform, {"status": U_STATUS_SUCCESS})
                current.platform_notes[platform].update(
                    {
                        "deal_prices_text": clean_text(deal_list_match.group("prices")),
                        "deal_count": clean_text(deal_list_match.group("count")),
                    }
                )
                continue

            summary_match = SUMMARY_RE.match(msg)
            if summary_match:
                value = to_float(summary_match.group("value"))
                label = summary_match.group("label")
                if label == U_QUOTE_AVG:
                    current.quote_avg = value
                elif label == U_DEAL_AVG:
                    current.deal_avg = value
                else:
                    current.final_price = value
                    current.final_price_logged = True
                continue

        if logger == "app.runtime":
            runtime_match = RUNTIME_RE.match(msg)
            if runtime_match:
                current.finished_at = ts
                current.elapsed_seconds = to_float(runtime_match.group("elapsed"))
                request = ast.literal_eval(runtime_match.group("request"))
                current.request_id = request.get("requestId")
                current.algorithm_mode = "DEFAULT"
                continue

    if current is not None:
        finalize_record(current)
        records.append(current)

    return records


def safe_sheet_name(name: str, used_names: Counter[str]) -> str:
    cleaned = "".join("_" if ch in INVALID_SHEET_CHARS else ch for ch in (name or "Sheet"))
    cleaned = cleaned[:31] or "Sheet"
    used_names[cleaned] += 1
    if used_names[cleaned] == 1:
        return cleaned
    suffix = f"_{used_names[cleaned]}"
    return cleaned[: 31 - len(suffix)] + suffix


def next_available_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    index = 1
    while True:
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def build_workbook(
    records: list[InquiryRecord],
    analysis_rows: Optional[list[AnalysisRow]] = None,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_title = PatternFill("solid", fgColor="A9D18E")
    fill_section = PatternFill("solid", fgColor="F4B183")
    fill_duplicate = PatternFill("solid", fgColor="E7E6E6")
    font_title = Font(bold=True, size=14)
    font_header = Font(bold=True)
    font_duplicate = Font(color="808080", italic=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    used_sheet_names: Counter[str] = Counter()

    def write_row(ws, row_index: int, values: list[object], header: bool = False) -> None:
        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if header:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = center
            else:
                cell.alignment = left
            cell.border = border

    def write_analysis_summary_sheet(ws, rows: list[AnalysisRow]) -> None:
        # 先展示最终偏差最大的记录，便于按偏差降序检查；无最终值的记录放最后。
        rows = sorted(
            rows,
            key=lambda row: (
                row.final_diff is None,
                -(row.final_diff or 0.0),
            ),
        )
        ws.merge_cells("A1:V1")
        title = ws["A1"]
        title.value = "抓取数据分析汇总"
        title.font = font_title
        title.fill = fill_title
        title.alignment = center
        title.border = border

        matched_count = sum(row.record is not None for row in rows)
        no_data_count = sum(row.evaluation_scope in {"无数据", "未匹配"} for row in rows)
        multi_count = sum(
            row.evaluation_scope in {"多峰排除", "多峰单独观察"}
            for row in rows
        )
        weak_reference_count = sum(bool(row.weak_reference_text) for row in rows)
        source_excluded_count = sum(
            row.evaluation_scope in {"原始数据排除", "多峰排除"}
            and not row.weak_reference_text
            for row in rows
        )
        algorithm_issue_count = sum(
            row.evaluation_scope == "算法问题" and not row.weak_reference_text
            for row in rows
        )
        within_count = sum(
            row.evaluation_scope == "计入单峰评价" and not row.weak_reference_text
            for row in rows
        )
        write_row(
            ws,
            3,
            [
                "总评估记录",
                len(rows),
                "日志匹配",
                matched_count,
                "无数据/未匹配",
                no_data_count,
                "多峰记录",
                multi_count,
                "原始数据排除",
                source_excluded_count,
                "算法问题",
                algorithm_issue_count,
                "单峰阈值内",
                within_count,
                "弱参考样本",
                weak_reference_count,
            ],
        )

        headers = [
            "Excel行号",
            "城市",
            "请求面积(㎡)",
            "小区",
            "评估单价",
            "选中主峰/原始主峰",
            "距评估最近候选峰偏差",
            "最终取值",
            "原始主峰偏差",
            "最终偏差",
            "决策分支",
            "所有候选峰（频率）",
            "分析结论",
            "评价口径",
        ]
        headers.extend(
            [
                "\u539f\u59cb\u623f\u6e90\u6570",
                "\u540c\u5e73\u53f0\u53bb\u91cd\u540e",
                "\u8de8\u5e73\u53f0\u53bb\u91cd\u540e",
                "\u8de8\u5e73\u53f0\u91cd\u590d\u7ec4",
            ]
        )
        headers.extend(
            [
                D_STRICT_CANDIDATES,
                D_WEAK_REFERENCE,
                "严格范围房源数",
                "弱参考补充数",
            ]
        )
        write_row(ws, 5, headers, header=True)

        for row_index, analysis in enumerate(rows, start=6):
            evaluation = analysis.evaluation
            record = analysis.record
            values = [
                evaluation.row_number,
                evaluation.city,
                evaluation.area,
                evaluation.community_name,
                evaluation.evaluation_price,
                analysis.raw_peak_price,
                analysis.nearest_peak_diff,
                record.final_price if record else None,
                analysis.raw_diff,
                analysis.final_diff,
                display_branch(record.branch_code) if record else "未匹配",
                analysis.candidate_text,
                analysis.conclusion,
                analysis.evaluation_scope,
                analysis.raw_listing_count,
                analysis.same_platform_listing_count,
                analysis.deduplicated_listing_count,
                analysis.cross_platform_duplicate_text,
                analysis.strict_candidate_text,
                analysis.weak_reference_text,
                analysis.strict_listing_count,
                analysis.weak_listing_count,
            ]
            write_row(ws, row_index, values)
            for column in (5, 6, 8):
                ws.cell(row=row_index, column=column).number_format = "#,##0.00"
            for column in (7, 9, 10):
                ws.cell(row=row_index, column=column).number_format = "0.00%"

        ws.freeze_panes = "A6"
        widths = {
            1: 10,
            2: 10,
            3: 14,
            4: 28,
            5: 14,
            6: 18,
            7: 18,
            8: 14,
            9: 14,
            10: 14,
            11: 28,
            12: 48,
            13: 34,
            14: 18,
            15: 14,
            16: 14,
            17: 14,
            18: 48,
            19: 48,
            20: 64,
            21: 14,
            22: 14,
        }
        for column, width in widths.items():
            ws.column_dimensions[get_column_letter(column)].width = width

    if analysis_rows is not None:
        summary_title = safe_sheet_name("分析汇总", used_sheet_names)
        summary_sheet = workbook.create_sheet(title=summary_title)
        write_analysis_summary_sheet(summary_sheet, analysis_rows)

    for record in records:
        ws = workbook.create_sheet(title=safe_sheet_name(record.community_name, used_sheet_names))
        ws.merge_cells("A1:J1")
        title = ws["A1"]
        title.value = record.community_name
        title.font = font_title
        title.fill = fill_title
        title.alignment = center
        title.border = border

        write_row(ws, 3, [U_CITY, U_REQUEST_AREA, D_REQUEST_ID, D_ALGORITHM_MODE], header=True)
        write_row(
            ws,
            4,
            [record.city, record.area, record.request_id or "", display_algorithm_mode(record.algorithm_mode)],
        )
        write_row(ws, 5, [D_STARTED, D_FINISHED, D_ELAPSED, D_RESULT], header=True)
        write_row(
            ws,
            6,
            [
                record.started_at,
                record.finished_at or "",
                record.elapsed_seconds,
                U_SUCCESS if record.success else "未完成",
            ],
        )
        write_row(ws, 7, [D_QUOTE, D_DEAL, D_FINAL, D_BRANCH], header=True)
        write_row(ws, 8, [record.quote_avg, record.deal_avg, record.final_price, display_branch(record.branch_code)])
        write_row(ws, 9, [D_BRANCH_TEXT, U_NOTE, "", ""], header=True)
        note_text = algorithm_description(record.algorithm_mode)
        if not record.success:
            if record.final_price_logged:
                if has_captured_data(record):
                    if record.algorithm_mode == "DEFAULT":
                        detail = weighted_median_candidates_detail(record)
                    else:
                        detail = calculation_no_data_description(record.algorithm_mode)
                    note_text = f"{note_text}；{detail}"
                else:
                    note_text = f"{note_text}；{U_NO_USABLE_DATA}。"
            else:
                note_text = f"{note_text}；{U_LOG_INCOMPLETE}"
        write_row(ws, 10, [record.branch_text or display_branch(record.branch_code), note_text, "", ""])
        current_candidates = (
            weighted_median_candidates_detail(record)
            if record.algorithm_mode == "DEFAULT"
            else ""
        )
        if record.algorithm_mode == "DEFAULT":
            deduplication = _listing_deduplication(record)
            same_platform_count = len(deduplication.same_platform_items)
            deduplicated_count = len(deduplication.items)
            current_candidates = (
                f"\u539f\u59cb\u623f\u6e90: {deduplication.raw_count} -> "
                f"\u540c\u5e73\u53f0\u53bb\u91cd: {same_platform_count} -> "
                f"\u8de8\u5e73\u53f0\u53bb\u91cd: {deduplicated_count}; "
                f"{current_candidates}"
            )
            if deduplication.cross_platform_groups:
                current_candidates += (
                    f"; \u8de8\u5e73\u53f0\u91cd\u590d\u7ec4: "
                    f"{_format_cross_platform_duplicate_text(deduplication)}"
                )
            weak_text = _format_final_weak_reference_text(record)
            if weak_text:
                current_candidates += f"; {D_WEAK_REFERENCE}: {weak_text}"
        write_row(ws, 11, [D_CANDIDATES, current_candidates, "", ""])

        ws.merge_cells("A12:J12")
        listing_section = ws["A12"]
        listing_section.value = U_LISTING_SECTION
        listing_section.font = font_header
        listing_section.fill = fill_section
        listing_section.alignment = left
        listing_section.border = border

        write_row(
            ws,
            13,
            [
                U_PLATFORM,
                U_STATUS,
                U_REASON,
                U_LISTING_COMMUNITY,
                U_TITLE,
                f"{U_AREA}({U_PINGMI_LABEL})",
                U_LAYOUT,
                f"{U_SELL}({U_PRICE_UNIT})",
                f"{U_TOTAL}({U_WAN})",
                U_DEDUP_STATUS,
            ],
            header=True,
        )

        dedup_statuses = _listing_dedup_statuses(record)
        grouped_listings: dict[str, list[ListingRow]] = defaultdict(list)
        for listing in record.listings:
            grouped_listings[listing.platform].append(listing)

        platforms = sorted(set(grouped_listings) | set(record.platform_notes))
        if not platforms:
            platforms = [""]

        row_index = 14
        for platform in platforms:
            note = record.platform_notes.get(platform, {})
            rows = grouped_listings.get(platform, [])
            if rows:
                for listing in rows:
                    write_row(
                        ws,
                        row_index,
                        [
                            platform,
                            display_status(note.get("status", U_STATUS_SUCCESS)),
                            _format_platform_reason(note),
                            listing.community_name,
                            listing.title,
                            listing.area,
                            listing.layout,
                            listing.unit_price,
                            listing.total_price,
                            dedup_statuses.get(id(listing), "\u4fdd\u7559\u7edf\u8ba1"),
                        ],
                    )
                    if id(listing) in dedup_statuses:
                        for column_index in range(1, 11):
                            cell = ws.cell(row=row_index, column=column_index)
                            cell.font = font_duplicate
                            cell.fill = fill_duplicate
                    row_index += 1
            else:
                write_row(
                    ws,
                    row_index,
                    [
                        platform,
                        display_status(note.get("status", "")),
                        _format_platform_reason(note),
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ],
                )
                row_index += 1

        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        deal_section = ws.cell(row=row_index, column=1, value=U_DEAL_SECTION)
        deal_section.font = font_header
        deal_section.fill = fill_section
        deal_section.alignment = left
        deal_section.border = border
        row_index += 1

        write_row(
            ws,
            row_index,
            [
                U_PLATFORM,
                U_STATUS,
                f"{U_AREA}({U_QM})",
                U_DATE,
                f"{U_TOTAL}({U_WAN})",
                f"{U_DANJIA}({U_PRICE_UNIT})",
                U_RAW,
            ],
            header=True,
        )
        row_index += 1

        deal_rows = 0
        for deal in record.deals:
            write_row(ws, row_index, [deal.platform, U_DEAL_RECORD, deal.area, deal.date, deal.total_price, deal.price, ""])
            row_index += 1
            deal_rows += 1

        for platform in platforms:
            note = record.platform_notes.get(platform, {})
            raw_parts = []
            if note.get("deal_source"):
                raw_parts.append(note["deal_source"])
                if note.get("deal_price"):
                    raw_parts.append(note["deal_price"])
            if note.get("deal_note"):
                raw_parts.append(note["deal_note"])
            if note.get("deal_prices_text"):
                raw_parts.append(note["deal_prices_text"])
            if raw_parts:
                write_row(ws, row_index, [platform, U_DEAL_NOTE, "", "", "", "", " / ".join(raw_parts)])
                row_index += 1
                deal_rows += 1

        if deal_rows == 0:
            write_row(ws, row_index, ["", U_DEAL_NOTE, "", "", "", "", U_NONE])
            row_index += 1

        ws.freeze_panes = "A13"
        for col_index, width in {
            1: 12,
            2: 12,
            3: 18,
            4: 14,
            5: 52,
            6: 12,
            7: 14,
            8: 14,
            9: 12,
            10: 28,
        }.items():
            ws.column_dimensions[get_column_letter(col_index)].width = width

    return workbook


def derive_output_path(
    log_path: Path,
    output_path: Optional[Path],
    evaluation_excel_path: Optional[Path] = None,
) -> Path:
    """在项目 results 目录下推导配套分析工作簿路径。"""
    if output_path is not None:
        return output_path
    project_root = Path(__file__).resolve().parents[2]
    results_dir = project_root / "results"
    source_stem = (
        evaluation_excel_path.stem
        if evaluation_excel_path is not None
        else log_path.stem
    )
    if not source_stem.endswith("_分析"):
        source_stem = f"{source_stem}_分析"
    return results_dir / f"{source_stem}.xlsx"


def save_workbook(workbook: Workbook, path: Path) -> Path:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path
    except PermissionError:
        fallback = next_available_path(path)
        fallback.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(fallback)
        return fallback


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("log_path", type=Path, help="Path to an inquiry log file")
    parser.add_argument(
        "-e",
        "--evaluation-excel",
        type=Path,
        help="Paired evaluation workbook; its stem is used for the XXX_分析.xlsx output name",
    )
    parser.add_argument("-o", "--output", type=Path, help="Output xlsx path")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.evaluation_excel is not None and not args.evaluation_excel.is_file():
        raise SystemExit(f"Evaluation workbook not found: {args.evaluation_excel}")
    lines = read_log_lines(args.log_path)
    records = parse_records(lines)
    if not records:
        raise SystemExit("No inquiry records parsed from log.")
    analysis_rows = None
    if args.evaluation_excel is not None:
        evaluation_rows = read_evaluation_rows(args.evaluation_excel)
        analysis_rows = analyze_evaluation_rows(evaluation_rows, records)
    workbook = build_workbook(records, analysis_rows)
    output_path = derive_output_path(
        args.log_path,
        args.output,
        args.evaluation_excel,
    )
    actual_path = save_workbook(workbook, output_path)
    complete_count = sum(
        1
        for record in records
        if record.final_price is not None or record.branch_code == "WEIGHTED_MEDIAN_MULTI"
    )
    incomplete_count = len(records) - complete_count
    total_listings = sum(len(record.listings) for record in records)
    total_deduplicated_listings = sum(
        len(_deduplicated_listing_rows(record)) for record in records
    )
    total_deals = sum(len(record.deals) for record in records)
    print(f"saved_to={actual_path}")
    print(f"records={len(records)}")
    print(f"complete={complete_count}")
    print(f"incomplete={incomplete_count}")
    print(f"listings={total_listings}")
    print(f"deduplicated_listings={total_deduplicated_listings}")
    print(f"deals={total_deals}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
