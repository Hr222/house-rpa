# -*- coding: utf-8 -*-
"""行舟深房 UI MVP 最小批量测试。

本脚本只复用 xzsfbj MVP，不启动全平台 API 服务，也不修改 MVP 主流程。
默认读取测试评估表的前 3 条有效记录，整批只启动一次 WMPF 桥，使用小程序
UI 完成搜索、面积筛选和在售 Network 采集；每条返回首页，整批完成后保持
桥接驻留等待用户主动关闭。探索时可传 --close-after-batch 提前清理。
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.algorithm import AlgorithmInput, evaluate_algorithm
from app.core.models import PlatformResult
from app.core.price_utils import round_price
from app.scripts import xzsfbj_mvp_test as xzsfbj_mvp


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_INPUT = PROJECT_ROOT / "test_data" / "房产评估汇总表_生成4_新增行政区.xlsx"
RESULT_DIR = PROJECT_ROOT / "results"
LOG_DIR = PROJECT_ROOT / "logs"


@dataclass(frozen=True)
class EvaluationItem:
    source_row: int
    city: str
    administrative_district: str
    community: str
    area: float
    evaluation_price: float


@dataclass
class BatchResult:
    item: EvaluationItem
    result: PlatformResult | None
    quote_avg: float | None = None
    final_price: float | None = None
    branch: str = "FAILED"
    elapsed_seconds: float | None = None
    error: str = ""


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="只运行行舟深房 UI MVP 的最小批量测试",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="评估表路径",
    )
    parser.add_argument(
        "--start",
        type=int,
        default=0,
        help="有效记录的 0-based 起始位置，默认 0",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=3,
        help="本次最多运行多少条，默认 3；全量可传 --limit 54",
    )
    parser.add_argument(
        "--gap",
        type=float,
        default=2.0,
        help="两条记录之间的等待秒数，默认 2 秒",
    )
    parser.add_argument(
        "--scroll-rounds",
        type=int,
        default=40,
        help="每条 UI MVP 真人滑动安全上限，默认 40",
    )
    parser.add_argument(
        "--ui-input",
        choices=("mouse", "touch"),
        default="mouse",
        help="UI 输入模型，默认 mouse",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="导出每条记录的 DOM、截图指纹和 Network 调试数据",
    )
    parser.add_argument(
        "--auto-ready",
        action="store_true",
        help="不等待首次人工回车；要求微信小程序已经停留在首页搜索状态",
    )
    parser.add_argument(
        "--close-after-batch",
        action="store_true",
        help="全部记录完成后关闭桥；默认保持桥接等待用户主动关闭",
    )
    parser.add_argument(
        "--continue-on-error",
        action="store_true",
        help="单条失败后继续下一条；默认失败即停止，避免页面状态串掉",
    )
    return parser


def _find_column(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def _read_items(path: Path) -> list[EvaluationItem]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if not header_values:
            raise ValueError(f"评估表没有表头: {path}")

        headers = {
            str(value).strip(): index
            for index, value in enumerate(header_values)
            if value is not None and str(value).strip()
        }
        city_col = _find_column(headers, "city", "城市")
        district_col = _find_column(headers, "行政区", "administrativeDistrict")
        area_col = _find_column(headers, "面积㎡", "面积")
        price_col = _find_column(headers, "评估单价")
        community_col = _find_column(headers, "小区名称", "小区")
        missing = [
            name
            for name, column in (
                ("行政区", district_col),
                ("面积㎡", area_col),
                ("评估单价", price_col),
                ("小区名称", community_col),
            )
            if column is None
        ]
        if missing:
            raise ValueError(f"评估表缺少列: {', '.join(missing)}")

        items: list[EvaluationItem] = []
        for source_row, values in enumerate(rows, start=2):
            community = values[community_col]
            area = values[area_col]
            evaluation_price = values[price_col]
            district = values[district_col]
            if not community or area is None or evaluation_price is None or not district:
                continue
            city = values[city_col] if city_col is not None else "深圳"
            items.append(
                EvaluationItem(
                    source_row=source_row,
                    city=str(city or "深圳").strip(),
                    administrative_district=str(district).strip(),
                    community=str(community).strip(),
                    area=float(area),
                    evaluation_price=float(evaluation_price),
                )
            )
        return items
    finally:
        workbook.close()


def _evaluate_result(result: PlatformResult) -> tuple[float | None, float | None, str]:
    if not result.quote_prices:
        return None, None, result.status.value
    evaluation = evaluate_algorithm(
        AlgorithmInput(
            quote_price_lists=[result.quote_prices],
            deal_price_lists=[],
        )
    )
    return (
        round_price(evaluation.quote_avg),
        round_price(evaluation.decision.final_price),
        evaluation.decision.branch,
    )


def _format_optional(value: float | None) -> str:
    return "" if value is None else f"{value:.2f}"


def _create_analysis_logger(path: Path) -> tuple[logging.Logger, logging.Handler]:
    path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("app.service")
    handler = logging.FileHandler(path, encoding="utf-8")
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(name)s - %(message)s")
    )
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    return logger, handler


def _write_analysis_record(logger: logging.Logger, batch: BatchResult) -> None:
    item = batch.item
    request = {
        "requestId": f"xzsfbj-batch-{item.source_row}",
        "city": item.city,
        "administrativeDistrict": item.administrative_district,
        "communityName": item.community,
        "area": item.area,
    }
    logger.info(
        "查询城市: %s, 小区: %s, 面积: %.2f㎡",
        item.city,
        item.community,
        item.area,
    )

    result = batch.result
    if result is not None:
        for snapshot in result.listing_snapshots:
            logger.info(
                "行舟深房: {小区名称: %s, 标题: %s, 面积: %s平米, 几房几厅: %s, "
                "售价: %s元/平, 总价: %s万, 房源编号: %s}",
                snapshot.community_name or "",
                snapshot.title or "",
                _format_optional(snapshot.area),
                snapshot.layout or "",
                _format_optional(snapshot.unit_price),
                _format_optional(snapshot.total_price),
                snapshot.house_id or "",
            )
        if batch.quote_avg is not None:
            logger.info("在售均价(单位:元/平): %.2f", batch.quote_avg)
        logger.info("最终取值(单位:元/平): %s", _format_optional(batch.final_price) or "None")
        logger.info("finalBranch: branchCode=%s", batch.branch)
        if not result.listing_snapshots:
            logger.info(
                "行舟深房: {状态: %s, 原因: %s}",
                result.status.value,
                result.reason or "无可用在售数据",
            )
    else:
        logger.info("行舟深房: {状态: ERROR, 原因: %s}", batch.error or "MVP 未产生结构化结果")
        logger.info("最终取值(单位:元/平): None")
        logger.info("finalBranch: branchCode=FAILED")

    logger.info(
        "[采集完成] request=%s 耗时=%.2f秒",
        request,
        batch.elapsed_seconds or 0.0,
    )


def _write_workbook(path: Path, results: list[BatchResult]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "结果"
    detail = workbook.create_sheet("在售明细")

    headers = [
        "源行号", "city", "行政区", "小区名称", "面积㎡", "评估单价",
        "状态", "在售条数", "在售均价", "最终取值", "差距比例%", "决策分支",
        "原因", "耗时秒",
    ]
    summary.append(headers)
    detail.append(["源行号", "小区名称", "房源编号", "小区名称", "面积㎡", "户型", "单价", "总价"])

    header_fill = PatternFill("solid", fgColor="4472C4")
    for sheet in (summary, detail):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    for batch in results:
        item = batch.item
        platform_result = batch.result
        status = platform_result.status.value if platform_result else "ERROR"
        listing_count = len(platform_result.listing_snapshots) if platform_result else 0
        diff = None
        if batch.final_price is not None and item.evaluation_price:
            diff = round((batch.final_price - item.evaluation_price) / item.evaluation_price * 100, 2)
        summary.append(
            [
                item.source_row,
                item.city,
                item.administrative_district,
                item.community,
                item.area,
                item.evaluation_price,
                status,
                listing_count,
                batch.quote_avg,
                batch.final_price,
                diff,
                batch.branch,
                (platform_result.reason if platform_result else batch.error) or "",
                batch.elapsed_seconds,
            ]
        )
        for snapshot in platform_result.listing_snapshots if platform_result else []:
            detail.append(
                [
                    item.source_row,
                    item.community,
                    snapshot.house_id,
                    snapshot.community_name,
                    snapshot.area,
                    snapshot.layout,
                    snapshot.unit_price,
                    snapshot.total_price,
                ]
            )

    valid_diffs = [
        row[10]
        for row in summary.iter_rows(min_row=2, values_only=True)
        if isinstance(row[10], (int, float))
    ]
    summary_row = summary.max_row + 2
    summary.cell(summary_row, 1, "批量统计")
    summary.cell(summary_row, 2, f"总记录: {len(results)}")
    summary.cell(summary_row + 1, 2, f"成功: {sum(1 for row in results if row.result is not None and row.result.quote_prices)}")
    summary.cell(summary_row + 2, 2, f"可比较偏差: {len(valid_diffs)}")
    if valid_diffs:
        summary.cell(summary_row + 3, 2, f"偏差≤10%: {sum(abs(value) <= 10 for value in valid_diffs)}")
        summary.cell(summary_row + 4, 2, f"平均偏差: {sum(valid_diffs) / len(valid_diffs):.2f}%")

    for sheet in (summary, detail):
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 36)
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


async def _run_batch(
    items: list[EvaluationItem],
    args: argparse.Namespace,
    on_complete=None,
) -> list[BatchResult]:
    xzsfbj_mvp.set_debug_mode(args.debug)
    targets = [
        (item.community, item.area, item.administrative_district)
        for item in items
    ]
    results: list[BatchResult] = []
    
    def consume_results(captured_results) -> None:
        results.clear()
        for index, item in enumerate(items):
            if index >= len(captured_results):
                break
            result, elapsed = captured_results[index]
            print(
                f"\n[{index + 1}/{len(items)}] {item.city} "
                f"{item.administrative_district} {item.community} "
                f"面积={item.area}㎡ 评估单价={item.evaluation_price}",
                flush=True,
            )
            if result is None:
                batch = BatchResult(
                    item=item,
                    result=None,
                    elapsed_seconds=elapsed,
                    error="UI MVP 批量流程异常中止，请检查上方日志",
                )
                results.append(batch)
                print(f"  失败: {batch.error}", flush=True)
                if not args.continue_on_error:
                    break
                continue

            quote_avg, final_price, branch = _evaluate_result(result)
            batch = BatchResult(
                item=item,
                result=result,
                quote_avg=quote_avg,
                final_price=final_price,
                branch=branch,
                elapsed_seconds=elapsed,
            )
            results.append(batch)
            print(
                f"  状态={result.status.value} 在售={len(result.listing_snapshots)} "
                f"最终取值={final_price} 分支={branch}",
                flush=True,
            )
        if on_complete is not None:
            on_complete(results)

    await xzsfbj_mvp.run_ui_mvp_batch(
        targets=targets,
        scroll_rounds=args.scroll_rounds,
        input_mode=args.ui_input,
        gap_seconds=args.gap,
        auto_ready=args.auto_ready,
        keep_alive=not args.close_after_batch,
        on_complete=consume_results,
    )
    return results


async def _async_main(args: argparse.Namespace) -> int:
    if args.start < 0 or args.limit <= 0:
        raise SystemExit("--start 必须 >= 0，--limit 必须 > 0")
    if args.scroll_rounds <= 0:
        raise SystemExit("--scroll-rounds 必须 > 0")

    all_items = _read_items(args.input)
    selected = all_items[args.start : args.start + args.limit]
    if not selected:
        raise SystemExit(f"没有可运行记录: 总有效记录={len(all_items)}, start={args.start}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_path = RESULT_DIR / f"行舟深房批量测试_{timestamp}.xlsx"
    log_path = LOG_DIR / f"xzsfbj_batch_{timestamp}-info.log"
    analysis_logger, handler = _create_analysis_logger(log_path)
    print(f"读取有效记录 {len(all_items)} 条，本次运行 {len(selected)} 条", flush=True)
    print(f"配套分析日志: {log_path}", flush=True)
    saved = False

    def save_results(results: list[BatchResult]) -> None:
        nonlocal saved
        for batch in results:
            _write_analysis_record(analysis_logger, batch)
        _write_workbook(result_path, results)
        saved = True
        print(f"\n结果已保存: {result_path}", flush=True)
        print(f"分析日志已保存: {log_path}", flush=True)
        print(
            "下一步可运行: "
            f"python .agents/skills/analyze-captured-data/scripts/export_operation_log_excel.py "
            f"{log_path} --evaluation-excel {result_path}",
            flush=True,
        )

    try:
        results = await _run_batch(selected, args, on_complete=save_results)
        if not saved:
            save_results(results)
    finally:
        handler.flush()
        handler.close()
        analysis_logger.removeHandler(handler)

    if not args.close_after_batch:
        print("批量结果已落盘，桥接继续驻留等待用户主动关闭；按 Ctrl+C 才会清理。", flush=True)
    return 0


def main() -> int:
    args = _build_parser().parse_args()
    return asyncio.run(_async_main(args))


if __name__ == "__main__":
    raise SystemExit(main())
