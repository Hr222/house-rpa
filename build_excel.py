# -*- coding: utf-8 -*-
"""从 operation.log 解析 app.service 打印的小区抓取 JSON，输出 Excel。

每条 `app.service - <平台>: {中文键: 值, ...}` 解析为一行，
列与 JSON 键一一对应，并补充「平台」「类型」两列标识来源。
"""
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

LOG_PATH = r"C:/Users/Administrator/Desktop/operation.log"
OUT_PATH = r"D:/workspace/rongzeyuan/jeethink-rpa/results/小区抓取数据_20260720.xlsx"

# 匹配: app.service - <平台标签>: <内容>
LINE_RE = re.compile(r"app\.service - (.+?): (.*)$")

PREFERRED_ORDER = ["小区名称", "标题", "面积", "几房几厅", "售价", "总价", "日期", "单价", "状态", "原因", "说明"]


def find_next_sep(s: str, start: int) -> int:
    """找到下一个真正的 'key: value' 分隔符位置（逗号+空格，且其后紧跟一个 key）。"""
    idx = start
    while True:
        c = s.find(", ", idx)
        if c == -1:
            return -1
        seg = s[c + 2:]
        colon2 = seg.find(":")
        comma2 = seg.find(",")
        if colon2 != -1 and (comma2 == -1 or colon2 < comma2):
            return c
        idx = c + 2


def parse_dict(s: str) -> dict:
    s = s.strip()
    # 去掉首尾可能存在的花括号（容忍行尾多余空白/嵌套）
    while s.startswith("{"):
        s = s[1:]
    while s.endswith("}"):
        s = s[:-1]
    s = s.strip()
    items = []
    i = 0
    n = len(s)
    while i < n:
        colon = s.find(":", i)
        if colon == -1:
            break
        key = s[i:colon].strip()
        j = colon + 1
        comma = find_next_sep(s, j)
        if comma == -1:
            val = s[j:].strip()
            items.append((key, val))
            break
        val = s[j:comma].strip()
        items.append((key, val))
        i = comma + 2
    return dict(items)


def main():
    with open(LOG_PATH, "r", encoding="utf-8") as f:
        lines = f.readlines()

    rows = []          # 每行一个 dict：含 平台/类型 + JSON 字段 / 说明
    all_keys = []      # JSON 键出现顺序（去重）

    for line in lines:
        m = LINE_RE.search(line)
        if not m:
            continue
        tag = m.group(1).strip()
        content = m.group(2).strip()

        # 仅处理平台数据行：内容以 { 开头(JJSON)，或标签以「成交」结尾(成败说明文字)
        is_deal_note = tag.endswith("成交") and not content.startswith("{")
        if not (content.startswith("{") or is_deal_note):
            continue

        if content.startswith("{"):
            d = parse_dict(content)
            if tag.endswith("成交"):
                platform, rtype = tag[:-2], "成交"
            else:
                platform, rtype = tag, ("无数据" if "状态" in d else "在售")
            rec = {"平台": platform, "类型": rtype}
            rec.update(d)
            for k in d.keys():
                if k not in all_keys:
                    all_keys.append(k)
        else:
            # 成交状态文字说明（无成交/未采集到/挂牌均价顶替等）
            rec = {"平台": tag[:-2], "类型": "成交", "说明": content}
        rows.append(rec)

    # 列顺序：标识列 + 按首选顺序排 JSON 键 + 其余
    cols = ["平台", "类型"]
    for k in PREFERRED_ORDER:
        if k in all_keys:
            cols.append(k)
            all_keys.remove(k)
    cols.extend(all_keys)  # 其余未知键追加在后面

    # 写 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "小区抓取数据"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rec in rows:
        ws.append([rec.get(col, "") for col in cols])

    # 列宽自适应（中文按 2 计）
    for c in range(1, len(cols) + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            length = sum(2 if ord(ch) > 127 else 1 for ch in s)
            max_len = max(max_len, length)
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max(max_len + 2, 8), 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(cols)).column_letter}{ws.max_row}"

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)

    # 控制台摘要
    print(f"总记录数: {len(rows)}")
    from collections import Counter
    cnt = Counter((r['平台'], r['类型']) for r in rows)
    for k, v in sorted(cnt.items()):
        print(f"  {k[0]} / {k[1]}: {v}")
    print("列顺序:", cols)
    print("输出文件:", OUT_PATH)


if __name__ == "__main__":
    main()
